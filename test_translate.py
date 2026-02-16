#!/usr/bin/env python3
"""
Test translation script for validating Excel files with the main translation logic.
Translates the first 3 rows of the selected file and prints the results for review.
"""
import os
import sys
import pandas as pd
from deep_translator import GoogleTranslator, LibreTranslator
import re
import string

def list_excel_files():
    return [f for f in os.listdir('.') if f.endswith('.xlsx')]

def choose_file(files, prompt):
    print(f"\n{prompt}")
    for idx, f in enumerate(files):
        print(f"  [{idx+1}] {f}")
    while True:
        choice = input("Enter number: ")
        if choice.isdigit() and 1 <= int(choice) <= len(files):
            return files[int(choice)-1]
        print("Invalid choice. Try again.")

def main():
    # Import mapping from language_mapping.py
    try:
        from language_mapping import LANGUAGE_MAPPING
    except ImportError:
        LANGUAGE_MAPPING = {}
    import pandas as pd
    from langdetect import detect, LangDetectException
    import difflib
    suspect_translations = []
    # Prompt for source language code
    default_source_lang = "en"
    source_lang = input(f"Enter source language code for the first column (default: {default_source_lang}): ").strip()
    if not source_lang:
        source_lang = default_source_lang
    print("\n=== Test Translation Script ===")
    excel_files = list_excel_files()
    if not excel_files:
        print("No .xlsx files found in current directory.")
        sys.exit(1)
    input_file = choose_file(excel_files, "Select input Excel file:")
    # Suggest default output name based on input
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    default_output_xlsx = f"{base_name}_test_results.xlsx"
    output_file_xlsx = input(f"Enter output Excel filename (default: {default_output_xlsx}): ").strip()
    if not output_file_xlsx:
        output_file_xlsx = default_output_xlsx
    if not output_file_xlsx.lower().endswith('.xlsx'):
        output_file_xlsx += '.xlsx'
    # Prompt before overwriting
    if os.path.exists(output_file_xlsx):
        confirm = input(f"Output file '{output_file_xlsx}' already exists. Overwrite? (y/n): ").strip().lower()
        if confirm != 'y':
            print("Aborting.")
            sys.exit(0)

    # --- [BOLD] row and multi-bold support ---
    df = pd.read_excel(input_file)
    bold_pairs = []
    rows_to_translate = []
    # Only process first 2 rows for translation
    for i in range(min(2, len(df))):
        cell_val = str(df.iloc[i, 0]) if pd.notna(df.iloc[i, 0]) else ""
        if cell_val.strip().startswith("[BOLD]"):
            if i > 0:
                bold_words = [w.strip() for w in cell_val.strip()[6:].split(",") if w.strip()]
                bold_pairs.append((i-1, bold_words, i))
            # Skip [BOLD] row from translation
        else:
            rows_to_translate.append(i)

    # Requirements check (minimal)
    try:
        import pandas, openpyxl, deep_translator
    except ImportError as e:
        print(f"Missing required package: {e.name}. Please install all dependencies with 'pip install -r requirements.txt'.")
        sys.exit(1)

    backend_options = ['1', '2', '3']
    print("\nChoose translation backend:")
    print("  [1] GoogleTranslator")
    print("  [2] LibreTranslator")
    print("  [3] Try Google, fallback to Libre")
    while True:
        backend_choice = input("Enter number: ")
        if backend_choice in backend_options:
            break
        print("Invalid choice. Try again.")
    # Ask user for terms to ignore (comma-separated, case-insensitive)
    ignore_terms_input = input("Enter comma-separated terms to ignore (leave blank for none): ").strip()
    if ignore_terms_input:
        ignore_terms = [t.strip() for t in ignore_terms_input.split(",") if t.strip()]
    else:
        ignore_terms = []
    # Preprocess input for clarity: clean whitespace, remove unnecessary punctuation, standardize casing
    def preprocess_text(text):
        text = ' '.join(str(text).split())
        allowed_punct = set('.!?,-:;')
        text = ''.join(ch for ch in text if ch.isalnum() or ch.isspace() or ch in allowed_punct)
        if text:
            text = text[0].upper() + text[1:]
        return text

    def extract_bold(text):
        bold_pattern = r"(\*\*([^*]+)\*\*)"
        placeholders = {}
        new_text = text
        bolds = re.findall(bold_pattern, new_text)
        for i, (full, inner) in enumerate(bolds):
            matched_ignore = None
            for term in ignore_terms:
                if inner.strip().lower() == term.lower():
                    matched_ignore = term
                    break
            if matched_ignore:
                placeholder = f"__IGNORE_{i}__"
                placeholders[placeholder] = f"[{matched_ignore}](#)"
                new_text = new_text.replace(full, placeholder, 1)
            else:
                placeholder = f"__BOLD_{i}__"
                placeholders[placeholder] = f"[{inner}](#)"
                new_text = new_text.replace(full, placeholder, 1)
        def ignore_replacer_factory(term):
            def ignore_replacer(match):
                idx = len([k for k in placeholders if k.startswith("__IGNORE_")])
                placeholder = f"__IGNORE_{idx}__"
                placeholders[placeholder] = f"[{term}](#)"
                return placeholder
            return ignore_replacer
        for term in ignore_terms:
            pattern = re.compile(re.escape(term), re.IGNORECASE)
            new_text = pattern.sub(ignore_replacer_factory(term), new_text)
        return new_text, placeholders
    def restore_bold(text, placeholders):
        for placeholder, link in placeholders.items():
            text = text.replace(placeholder, link)
        return text

    # Warn if first column is empty
    if df.shape[1] == 0 or df.iloc[:,0].isnull().all() or (df.iloc[:,0].astype(str).str.strip() == '').all():
        print("Warning: The first column (source text) is empty. Please check your input file.")
        sys.exit(1)
    for col in df.columns:
        df[col] = df[col].astype('object')
    # Validate columns: support EITHER all codes or all mapped names (not both)
    col_headers = [str(col).strip() for col in df.columns[1:]]
    print("\n[DEBUG] Input column headers:", list(df.columns))
    print("[DEBUG] Processed col_headers:", col_headers)
    try:
        code_dict = GoogleTranslator().get_supported_languages(as_dict=True)
        supported_codes = set(code_dict.values())
    except Exception:
        try:
            supported_codes = set(GoogleTranslator().get_supported_languages())
        except Exception:
            supported_codes = set()
    valid_columns = []  # List of (col_name, target_code)
    skipped_codes = []
    for col in col_headers:
        if not col:
            skipped_codes.append("<empty header>")
        elif col in supported_codes:
            valid_columns.append((col, col))
        elif col in LANGUAGE_MAPPING:
            valid_columns.append((col, LANGUAGE_MAPPING[col]))
        else:
            skipped_codes.append(col)
    if skipped_codes:
        print(f"Info: The following columns are intentionally skipped (unsupported or empty): {', '.join(map(str, skipped_codes))}")

    # Prepare exclusion report for summary
    exclusion_report = ""
    if skipped_codes:
        exclusion_report = f"Excluded columns (unsupported or empty): {', '.join(map(str, skipped_codes))}"
    print("\nTesting translation on first 3 rows...")
    # Only keep the source and valid language columns in the test output
    # Build test_df with all valid columns (even duplicates), skipping only invalid/empty
    # Output columns should always match input headers (or mapped names), never auto-corrected
    test_df = df.loc[:1, list(df.columns)].copy()  # Only first 2 rows
    success_count = 0
    fail_count = 0
    failed_translations = []
    # --- Context-aware [BOLD] handling for test rows ---
    context_bold_rows = []
    try:
        # Loop over valid columns only
        for col_name, target_code in valid_columns:
            lang_code = col_name
            print(f"Translating column: {col_name} (using code: {target_code})")
            print(f"[DEBUG] lang_code: {lang_code}, target_code: {target_code}")
            for row_idx in rows_to_translate:
                source_text = str(df.iat[row_idx, 0]).strip()
                prepped_text = preprocess_text(source_text)
                print(f"Row {row_idx+1} ({source_lang.upper()}): {source_text}")
                max_attempts = 3
                attempt = 0
                error = None
                translated = None
                while attempt < max_attempts:
                    try:
                        if backend_choice == '1':
                            translated = GoogleTranslator(source=source_lang, target=target_code).translate(prepped_text)
                            error = None
                        elif backend_choice == '2':
                            translated = LibreTranslator(source=source_lang, target=target_code).translate(prepped_text)
                            error = None
                        else:
                            translated = GoogleTranslator(source=source_lang, target=target_code).translate(prepped_text)
                            error = None
                        break
                    except Exception as e:
                        error = e
                        attempt += 1
                if error:
                    test_df.at[row_idx, col_name] = f"[ERROR: {error}]"
                    print(f"  {lang_code}: [ERROR: {error}]")
                    fail_count += 1
                    failed_translations.append({
                        "row": row_idx,
                        "english_text": source_text,
                        "language_code": lang_code,
                        "error": str(error)
                    })
                else:
                    translated_str = str(translated)
                    test_df.at[row_idx, col_name] = translated_str
                    print(f"  {lang_code}: {translated_str}")
                    success_count += 1
            # --- Context-aware [BOLD] handling for test rows ---
            for (main_idx, bold_words, bold_row_idx) in bold_pairs:
                if main_idx == row_idx:
                    bold_translations = []
                    for bold_word in bold_words:
                        try:
                            bold_translated = GoogleTranslator(source=source_lang, target=target_code).translate(bold_word)
                        except Exception:
                            bold_translated = ""
                        found_in_sentence = False
                        if bold_translated and bold_translated in translated_str:
                            found_in_sentence = True
                            bold_translations.append(f"{bold_word} → {bold_translated} (in sentence)")
                        else:
                            import difflib
                            matches = difflib.get_close_matches(bold_translated, translated_str.split(), n=1, cutoff=0.7)
                            if matches:
                                found_in_sentence = True
                                bold_translations.append(f"{bold_word} → {matches[0]} (fuzzy match)")
                            else:
                                bold_translations.append(f"{bold_word} → {bold_translated} (not found)")
                    context_bold_rows.append({
                        "Row": row_idx+2,
                        "Language": lang_code,
                        "Source": source_text,
                        "Bold Words & Translations": "; ".join(bold_translations)
                    })
            for row_idx in range(test_df.shape[0]):
                source_text = str(test_df.iat[row_idx, 0]).strip()
                prepped_text = preprocess_text(source_text)
                print(f"Row {row_idx+1} ({source_lang.upper()}): {source_text}")
                max_attempts = 3
                attempt = 0
                error = None
                translated = None
                while attempt < max_attempts:
                    try:
                        if backend_choice == '1':
                            translated = GoogleTranslator(source=source_lang, target=target_code).translate(prepped_text)
                            error = None
                        elif backend_choice == '2':
                            translated = LibreTranslator(source=source_lang, target=target_code).translate(prepped_text)
                            error = None
                        else:
                            translated = GoogleTranslator(source=source_lang, target=target_code).translate(prepped_text)
                            error = None
                        break
                    except Exception as e:
                        error = e
                        attempt += 1
                if error:
                    test_df.at[row_idx, col_name] = f"[ERROR: {error}]"
                    print(f"  {lang_code}: [ERROR: {error}]")
                    fail_count += 1
                    failed_translations.append({
                        "row": row_idx,
                        "english_text": source_text,
                        "language_code": lang_code,
                        "error": str(error)
                    })
                else:
                    translated_str = str(translated)
                    test_df.at[row_idx, col_name] = translated_str
                    print(f"  {lang_code}: {translated_str}")
                    success_count += 1

                    # --- Context-aware [BOLD] handling for test rows ---
                    for (main_idx, bold_words, bold_row_idx) in bold_pairs:
                        if main_idx == row_idx:
                            bold_translations = []
                            for bold_word in bold_words:
                                try:
                                    bold_translated = GoogleTranslator(source=source_lang, target=target_code).translate(bold_word)
                                except Exception:
                                    bold_translated = ""
                                found_in_sentence = False
                                if bold_translated and bold_translated in translated_str:
                                    found_in_sentence = True
                                    bold_translations.append(f"{bold_word} → {bold_translated} (in sentence)")
                                else:
                                    import difflib
                                    matches = difflib.get_close_matches(bold_translated, translated_str.split(), n=1, cutoff=0.7)
                                    if matches:
                                        found_in_sentence = True
                                        bold_translations.append(f"{bold_word} → {matches[0]} (fuzzy match)")
                                    else:
                                        if bold_word in ignore_terms:
                                            bold_translations.append(f"{bold_word} → {bold_translated} (not found in main text; IGNORED TERM, informational only)")
                                        else:
                                            bold_translations.append(f"{bold_word} → {bold_translated} (not found in main text; this is informational, not an error)")
                            context_bold_rows.append({
                                "Row": row_idx+2,
                                "Language": lang_code,
                                "Source": source_text,
                                "Bold Words & Translations": "; ".join(bold_translations)
                            })

                    # --- Translation verification: back-translate and language detect ---
                    try:
                        back_translated = GoogleTranslator(source=target_code, target=source_lang).translate(translated_str)
                        similarity = difflib.SequenceMatcher(None, source_text, back_translated).ratio() if back_translated else 0.0
                        try:
                            detected_lang = detect(translated_str)
                        except LangDetectException:
                            detected_lang = "unknown"
                        if similarity < 0.8 or (
                            detected_lang != target_code and
                            detected_lang != lang_code and
                            detected_lang != "unknown"
                        ):
                            suspect_translations.append({
                                "row": row_idx,
                                "source_text": source_text,
                                "language_code": lang_code,
                                "translated_text": translated_str,
                                "back_translated": back_translated,
                                "similarity": similarity,
                                "detected_lang": detected_lang
                            })
                    except Exception:
                        pass
            print(f"Finished translating column: {lang_code}")
    except KeyboardInterrupt:
        print("\nTest interrupted by user.")
        print(f"Rows translated: {success_count}, Failed: {fail_count}")
        sys.exit(1)

    # Output to user-specified files
    test_df.to_excel(output_file_xlsx, index=False)
        # CSV output has been removed as per the update.

    # Output context-aware bold translations to a separate sheet

    # Save test results to Excel and CSV, ensuring only valid columns are included
    valid_language_cols = [col for col, _ in valid_columns]
    output_cols = [df.columns[0]] + valid_language_cols
    # Insert [BOLD] rows as new rows in the main sheet after each main row (always)
    new_rows = []
    for idx in range(test_df.shape[0]):
        new_rows.append(test_df.iloc[idx])
        # Check if this row has a [BOLD] pair
        for (main_idx, bold_words, bold_row_idx) in bold_pairs:
            if main_idx == idx:
                if bold_row_idx == idx:
                    # Build a new row for bold words, keep '[BOLD]' in the source column for clarity
                    phrase = test_df.iloc[idx, 0].strip()
                    if not phrase.startswith('[BOLD]'):
                        phrase = '[BOLD] ' + phrase
                    bold_row = pd.Series([phrase], index=[test_df.columns[0]])
                    for col_name, target_code in valid_columns:
                        translated_bolds = []
                        for bold_word in bold_words:
                            try:
                                bold_translated = GoogleTranslator(source=source_lang, target=target_code).translate(bold_word)
                            except Exception:
                                bold_translated = ""
                            if not bold_translated or bold_translated.strip() == "":
                                # Italicize if not translated
                                translated_bolds.append(f"*{bold_word}*")
                            else:
                                translated_bolds.append(str(bold_translated))
                        bold_row[col_name] = ", ".join(translated_bolds)
                    new_rows.append(bold_row)
    test_df_with_bold = pd.DataFrame(new_rows, columns=test_df.columns)
    test_df_with_bold[output_cols].to_excel(output_file_xlsx, index=False, sheet_name="Translations")

    # --- Apply real bold formatting in Excel for markdown bold (**...**) ---
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.load_workbook(output_file_xlsx)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):  # skip header
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "**" in cell.value:
                import re
                matches = list(re.finditer(r"\*\*([^*]+)\*\*", cell.value))
    summary_report = [
        "Summary:",
        f"  Successful translations: {success_count}",
        f"  Failed translations: {fail_count}",
        f"  Total cells tested: {success_count + fail_count}"
    ]
    if exclusion_report:
        summary_report.append("")
        summary_report.append(exclusion_report)
    if failed_translations:
        summary_report.append(f"First 3 failed translations:")
        for fail in failed_translations[:3]:
            # Output columns should always match input headers, never auto-corrected or mapped
            test_df = df.loc[:1, list(df.columns)].copy()  # Only first 2 rows
        summary_report.append("")
        summary_report.append(f"Suspect translations flagged: {len(suspect_translations)}")
        for s in suspect_translations[:3]:
            summary_report.append(f"  Row {s['row']+1}, Language: {s['language_code']}, Similarity: {s['similarity']:.2f}, Detected: {s['detected_lang']}")
    with open("test_translation_summary_report.txt", "w", encoding="utf-8") as summary_file:
        summary_file.write("\n".join(summary_report))
    print("Summary report saved to test_translation_summary_report.txt.")

if __name__ == "__main__":
    main()

# --- Utility: Detect and print bold cells in Translations sheet ---
def print_bold_cells_in_excel(filename):
    import openpyxl
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    print(f"Bold cells in '{filename}' (active sheet):")
    for row in ws.iter_rows(min_row=2):  # skip header
        for cell in row:
            if cell.font and cell.font.bold:
                print(f"  Cell {cell.coordinate}: {cell.value}")
