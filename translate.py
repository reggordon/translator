

#!/usr/bin/env python3
# USAGE: Make this file executable with 'chmod +x translate.py' and run with './translate.py'


import os
import sys
import time
import csv
import threading
import pandas as pd
from deep_translator import GoogleTranslator, LibreTranslator
from tqdm import tqdm
from langdetect import detect, LangDetectException
import difflib
import openpyxl
from openpyxl.styles import PatternFill



# List all Excel files in the current directory
def list_excel_files():
    return [f for f in os.listdir('.') if f.endswith('.xlsx')]


# Prompt user to select a file from a list
def choose_file(files, prompt):
    print(f"\n{prompt}")
    for idx, f in enumerate(files):
        print(f"  [{idx+1}] {f}")
    while True:
        choice = input("Enter number: ")
        if choice.isdigit() and 1 <= int(choice) <= len(files):
            return files[int(choice)-1]
        print("Invalid choice. Try again.")


# Prompt user to select translation backend
def choose_backend():
    print("\nChoose translation backend:")
    print("  [1] GoogleTranslator")
    print("  [2] LibreTranslator")
    print("  [3] Try Google, fallback to Libre")
    while True:
        choice = input("Enter number: ")
        if choice in ['1', '2', '3']:
            return choice
        print("Invalid choice. Try again.")


def main():
    # Prompt for source language code
    default_source_lang = "en"
    source_lang = input(f"Enter source language code for the first column (default: {default_source_lang}): ").strip()
    if not source_lang:
        source_lang = default_source_lang
    print("\n=== Translation Script ===")
    excel_files = list_excel_files()
    if not excel_files:
        print("No .xlsx files found in current directory.")
        sys.exit(1)



    input_file = choose_file(excel_files, "Select input Excel file:")
    # Suggest default output name based on input
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    default_output = f"{base_name}_translated.xlsx"
    output_file = input(f"Enter output Excel filename (default: {default_output}): ").strip()
    if not output_file:
        output_file = default_output
    if not output_file.lower().endswith('.xlsx'):
        output_file += '.xlsx'
    # Prompt before overwriting
    if os.path.exists(output_file):
        confirm = input(f"Output file '{output_file}' already exists. Overwrite? (y/n): ").strip().lower()
        if confirm != 'y':
            print("Aborting.")
            sys.exit(0)

    # Requirements check (minimal)
    try:
        import pandas, openpyxl, deep_translator, tqdm
    except ImportError as e:
        print(f"Missing required package: {e.name}. Please install all dependencies with 'pip install -r requirements.txt'.")
        sys.exit(1)

    # Ask user for terms to ignore (comma-separated, case-insensitive)
    ignore_terms_input = input("Enter comma-separated terms to ignore (leave blank for none): ").strip()
    if ignore_terms_input:
        ignore_terms = [t.strip() for t in ignore_terms_input.split(",") if t.strip()]
    else:
        ignore_terms = []

    import re


    df = pd.read_excel(input_file, dtype=str)

    # Detect [BOLD] rows and build a list of (main_row_idx, bold_word) pairs
    bold_pairs = []
    rows_to_translate = []
    i = 0
    while i < len(df):
        cell_val = str(df.iloc[i, 0]) if pd.notna(df.iloc[i, 0]) else ""
        if cell_val.strip().startswith("[BOLD]"):
            # This is a [BOLD] row, pair with previous row
            if i > 0:
                bold_words = [w.strip() for w in cell_val.strip()[6:].split(",") if w.strip()]
                bold_pairs.append((i-1, bold_words, i))
            i += 1
        else:
            rows_to_translate.append(i)
            i += 1

    # Warn if first column is empty
    if df.shape[1] == 0 or df.iloc[:,0].isnull().all() or (df.iloc[:,0].astype(str).str.strip() == '').all():
        print("Warning: The first column (source text) is empty. Please check your input file.")
        sys.exit(1)

    # Ensure all columns are of type 'object' to avoid dtype issues when assigning strings
    for col in df.columns:
        df[col] = df[col].astype('object')

    # Validate language codes: only keep those supported by GoogleTranslator or LibreTranslator

    try:
        google_dict = GoogleTranslator().get_supported_languages(as_dict=True)
        google_codes = set(google_dict.values())
    except Exception:
        google_codes = set()
    try:
        libre_dict = LibreTranslator().get_supported_languages(as_dict=True)
        libre_codes = set(libre_dict.values())
    except Exception:
        libre_codes = set()
    supported_codes = google_codes | libre_codes
    language_codes = []
    skipped_codes = []
    for col in df.columns[1:]:
        col_str = str(col).strip()
        if not col_str:
            skipped_codes.append("<empty header>")
            continue
        # Use the full code for validation
        if col_str in supported_codes:
            language_codes.append(col)
        else:
            skipped_codes.append(col_str)
    if skipped_codes:
        print(f"Warning: Skipping unsupported or empty language columns: {', '.join(map(str, skipped_codes))}")
        # Drop problematic columns from df to keep headers/data aligned
        cols_to_drop = [col for col in df.columns[1:] if str(col).strip() in skipped_codes or not str(col).strip()]
        if cols_to_drop:
            df.drop(columns=cols_to_drop, inplace=True)

    # Prepare exclusion report for summary
    exclusion_report = ""
    if skipped_codes:
        exclusion_report = f"Excluded columns (unsupported or empty): {', '.join(map(str, skipped_codes))}"

    # Check for empty language columns and alert the user
    empty_lang_cols = []
    for col in language_codes:
        if df[col].isnull().all() or (df[col].astype(str).str.strip() == '').all():
            empty_lang_cols.append(col)
    if empty_lang_cols:
        print(f"Note: The following language columns are completely empty and will be filled by translation: {', '.join(map(str, empty_lang_cols))}")


    backend_choice = choose_backend()

    failed_translations = []
    # rows_to_translate now only includes main text rows (not [BOLD] rows)

    success_count = 0
    skip_count = 0
    fail_count = 0
    suspect_translations = []



    # Preprocess input for clarity: clean whitespace, remove unnecessary punctuation, standardize casing
    import string
    def preprocess_text(text):
        # Remove extra whitespace
        text = ' '.join(str(text).split())
        # Remove unnecessary punctuation except for basic sentence structure
        allowed_punct = set('.!?,-:;')
        text = ''.join(ch for ch in text if ch.isalnum() or ch.isspace() or ch in allowed_punct)
        # Standardize casing (capitalize first letter, rest lower)
        if text:
            text = text[0].upper() + text[1:]
        return text

    # Helper: extract bold markdown (**) and replace with placeholders
    # Also, always preserve user-specified ignore terms (case-insensitive) as links and never translate them
    def extract_bold(text):
        bold_pattern = r"(\*\*([^*]+)\*\*)"
        placeholders = {}
        new_text = text
        # First, handle bolds
        bolds = re.findall(bold_pattern, new_text)
        for i, (full, inner) in enumerate(bolds):
            # If the bolded text matches any ignore term, always treat as link and preserve
            matched_ignore = None
            for term in ignore_terms:
                if inner.strip().lower() == term.lower():
                    matched_ignore = term
                    break
            if matched_ignore:
                placeholder = f"__IGNORE_{i}__"
                placeholders[placeholder] = f"[{matched_ignore}](#)"
                new_text = new_text.replace(full, placeholder, 1)
            else:
                placeholder = f"__BOLD_{i}__"
                placeholders[placeholder] = f"[{inner}](#)"
                new_text = new_text.replace(full, placeholder, 1)
        # Then, handle any remaining ignore terms (not bolded)
        def ignore_replacer_factory(term):
            def ignore_replacer(match):
                idx = len([k for k in placeholders if k.startswith("__IGNORE_")])
                placeholder = f"__IGNORE_{idx}__"
                placeholders[placeholder] = f"[{term}](#)"
                return placeholder
            return ignore_replacer
        for term in ignore_terms:
            pattern = re.compile(re.escape(term), re.IGNORECASE)
            new_text = pattern.sub(ignore_replacer_factory(term), new_text)
        return new_text, placeholders

    def restore_bold(text, placeholders):
        for placeholder, link in placeholders.items():
            text = text.replace(placeholder, link)
        return text

    # Helper: run translation with timeout
    def translate_with_timeout(func, args=(), timeout=15):
        result = {}
        def wrapper():
            try:
                result['value'] = func(*args)
            except Exception as e:
                result['error'] = e
        thread = threading.Thread(target=wrapper)
        thread.start()
        thread.join(timeout)
        if thread.is_alive():
            return None, TimeoutError('Translation timed out')
        return result.get('value'), result.get('error')


    try:
        # Store context-aware bold translations for output
        context_bold_rows = []
        for col_idx, lang_code in enumerate(language_codes, start=1):
            print(f"Translating column: {lang_code}")
            target_code = str(lang_code).strip()
            for row_idx in rows_to_translate:
                english_text = str(df.iat[row_idx, 0]).strip()
                prepped_text = preprocess_text(english_text)
                if pd.notna(df.iat[row_idx, col_idx]) and str(df.iat[row_idx, col_idx]).strip() != "":
                    skip_count += 1
                    continue
                try:
                    translated = None
                    backend = ""
                    max_attempts = 3
                    attempt = 0
                    error = None
                    while attempt < max_attempts:
                        if backend_choice == '1':
                            translated, error = translate_with_timeout(
                                GoogleTranslator(source=source_lang, target=target_code).translate,
                                (prepped_text,), 15)
                            backend = "Google"
                        elif backend_choice == '2':
                            translated, error = translate_with_timeout(
                                LibreTranslator(source=source_lang, target=target_code).translate,
                                (prepped_text,), 15)
                            backend = "Libre"
                        else:
                            translated, error = translate_with_timeout(
                                GoogleTranslator(source=source_lang, target=target_code).translate,
                                (prepped_text,), 15)
                            backend = "Google"
                            if error:
                                translated, error = translate_with_timeout(
                                    LibreTranslator(source=source_lang, target=target_code).translate,
                                    (prepped_text,), 15)
                                backend = "Libre"
                        if not error:
                            break
                        attempt += 1
                        time.sleep(0.5)
                    if error:
                        raise error
                    translated_str = str(translated)
                    df.iat[row_idx, col_idx] = translated_str
                    success_count += 1

                    # Context-aware [BOLD] handling
                    for (main_idx, bold_words, bold_row_idx) in bold_pairs:
                        if main_idx == row_idx:
                            bold_translations = []
                            for bold_word in bold_words:
                                # Check if bold_word is in main text
                                if bold_word not in english_text:
                                    print(f"[WARN] [BOLD] word '{bold_word}' not found in main text at row {row_idx+2}")
                                # Try to find translation of bold_word in translated_str
                                try:
                                    bold_translated, _ = translate_with_timeout(
                                        GoogleTranslator(source=source_lang, target=target_code).translate,
                                        (bold_word,), 15)
                                except Exception:
                                    bold_translated = ""
                                found_in_sentence = False
                                if bold_translated and bold_translated in translated_str:
                                    found_in_sentence = True
                                    bold_translations.append(f"{bold_word} → {bold_translated} (in sentence)")
                                else:
                                    import difflib
                                    matches = difflib.get_close_matches(bold_translated, translated_str.split(), n=1, cutoff=0.7)
                                    if matches:
                                        found_in_sentence = True
                                        bold_translations.append(f"{bold_word} → {matches[0]} (fuzzy match)")
                                    else:
                                        bold_translations.append(f"{bold_word} → {bold_translated} (not found)")
                            context_bold_rows.append({
                                "Row": row_idx+2,
                                "Language": lang_code,
                                "Source": english_text,
                                "Bold Words & Translations": "; ".join(bold_translations)
                            })

                    try:
                        back_translated, bt_error = translate_with_timeout(
                            GoogleTranslator(source=target_code, target=source_lang).translate,
                            (translated_str,), 15)
                        if bt_error:
                            back_translated = ""
                        similarity = difflib.SequenceMatcher(None, english_text, back_translated).ratio() if back_translated else 0.0
                        try:
                            detected_lang = detect(translated_str)
                        except LangDetectException:
                            detected_lang = "unknown"
                        # Technical document: more forgiving criteria
                        if similarity < 0.7 or (
                            detected_lang not in [target_code, lang_code, "unknown", "en"]
                        ):
                            suspect_translations.append({
                                "row": row_idx,
                                "english_text": english_text,
                                "language_code": lang_code,
                                "short_code": target_code,
                                "translated_text": translated_str,
                                "back_translated": back_translated,
                                "similarity": similarity,
                                "detected_lang": detected_lang
                            })
                    except Exception:
                        pass

                except Exception as e:
                    failed_translations.append({
                        "row": row_idx,
                        "english_text": english_text,
                        "language_code": lang_code,
                        "short_code": target_code,
                        "error": str(e)
                    })
                    df.iat[row_idx, col_idx] = ""
                    backend = "FAILED"
                    fail_count += 1
                time.sleep(0.3)
            print(f"Finished translating column: {lang_code}")
    except KeyboardInterrupt:
        print("\nTranslation interrupted by user.")
        print(f"Rows translated: {success_count}, Failed: {fail_count}, Skipped: {skip_count}")
        sys.exit(1)

    # Save main results and suspects to separate sheets in the same Excel file
    import openpyxl
    from pandas import ExcelWriter

    # Ensure output file is created even if there are no translations

    with ExcelWriter(output_file, engine="openpyxl") as writer:
        # Only keep the source and valid language columns in the output
        output_cols = [df.columns[0]] + language_codes
        if df.shape[0] == 0:
            empty_df = pd.DataFrame(columns=output_cols)
            empty_df.to_excel(writer, index=False, sheet_name="Translations")
        else:
            df[output_cols].to_excel(writer, index=False, sheet_name="Translations")
        if suspect_translations:
            suspects_df = pd.DataFrame(suspect_translations)
            suspects_df.to_excel(writer, index=False, sheet_name="SuspectTranslations")

        # Output context-aware bold translations
        if context_bold_rows:
            context_bold_df = pd.DataFrame(context_bold_rows)
            context_bold_df.to_excel(writer, index=False, sheet_name="ContextBoldWords")

    # --- Apply real bold formatting in Excel for markdown bold (**...**) ---
    import re
    from openpyxl.styles import Font
    wb = openpyxl.load_workbook(output_file)
    for sheet_name in ["Translations", "SuspectTranslations"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2):  # skip header
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "**" in cell.value:
                        # Find all bold regions
                        matches = list(re.finditer(r"\*\*([^*]+)\*\*", cell.value))
                        if matches:
                            # Remove markdown and apply bold to those regions
                            new_val = re.sub(r"\*\*([^*]+)\*\*", r"\1", cell.value)
                            cell.value = new_val
                            # Apply rich text bold if possible (Excel supports only for entire cell in openpyxl)
                            cell.font = Font(bold=True)
    wb.save(output_file)

    # Highlight suspect translations in the main Translations sheet
    if suspect_translations:
        wb = openpyxl.load_workbook(output_file)
        ws = wb["Translations"]
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # Build a set of (row, col) for suspect translations (1-based for openpyxl)
        suspect_cells = set()
        output_cols = [df.columns[0]] + language_codes
        for s in suspect_translations:
            excel_row = s["row"] + 2  # +1 for header, +1 for 0-based index
            try:
                excel_col = output_cols.index(s["language_code"]) + 1
            except ValueError:
                continue
            suspect_cells.add((excel_row, excel_col))
        for row, col in suspect_cells:
            ws.cell(row=row, column=col).fill = yellow_fill
        wb.save(output_file)

    # Still save failures to CSV for easy review
    if failed_translations:
        with open("failed_translations_log.csv", mode="w", newline="", encoding="utf-8") as log_file:
            writer = csv.DictWriter(
                log_file,
                fieldnames=["row", "english_text", "language_code", "short_code", "error"]
            )
            writer.writeheader()
            writer.writerows(failed_translations)


    print(f"\n✅ Translations complete. Results saved to '{output_file}'.")
    print(f"Summary:")
    print(f"  Successful translations: {success_count}")
    print(f"  Skipped cells (already translated): {skip_count}")
    print(f"  Failed translations: {fail_count}")
    print(f"  Suspect translations (review): {len(suspect_translations)}")
    if failed_translations:
        print(f"⚠️ {len(failed_translations)} failures logged to 'failed_translations_log.csv'")
        print("First 3 failed translations:")
        for fail in failed_translations[:3]:
            print(f"  Row {fail['row']+1}, Language: {fail['language_code']}, Error: {fail['error']}")
    if suspect_translations:
        print(f"⚠️ {len(suspect_translations)} suspect translations are included as a sheet in the output Excel file.")

    # Save summary report to file
    summary_report = [
        f"Translations complete. Results saved to '{output_file}'.",
        "Summary:",
        f"  Successful translations: {success_count}",
        f"  Skipped cells (already translated): {skip_count}",
        f"  Failed translations: {fail_count}",
        f"  Suspect translations (review): {len(suspect_translations)}"
    ]
    if exclusion_report:
        summary_report.append("")
        summary_report.append(exclusion_report)
    if failed_translations:
        summary_report.append(f"{len(failed_translations)} failures logged to 'failed_translations_log.csv'")
        for fail in failed_translations[:3]:
            summary_report.append(f"  Row {fail['row']+1}, Language: {fail['language_code']}, Error: {fail['error']}")
    if suspect_translations:
        summary_report.append(f"{len(suspect_translations)} suspect translations are included as a sheet ('SuspectTranslations') in the output Excel file.")
    with open("translation_summary_report.txt", "w", encoding="utf-8") as summary_file:
        summary_file.write("\n".join(summary_report))


if __name__ == "__main__":
    main()

# --- Utility: Detect and print bold cells in Translations sheet ---
def print_bold_cells_in_excel(filename):
    import openpyxl
    wb = openpyxl.load_workbook(filename)
    if "Translations" not in wb.sheetnames:
        print("No 'Translations' sheet found.")
        return
    ws = wb["Translations"]
    print(f"Bold cells in '{filename}' (Translations sheet):")
    for row in ws.iter_rows(min_row=2):  # skip header
        for cell in row:
            if cell.font and cell.font.bold:
                print(f"  Cell {cell.coordinate}: {cell.value}")
